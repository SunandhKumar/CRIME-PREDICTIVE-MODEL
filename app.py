import os, sys, shutil, time

from flask import Flask, request, jsonify, render_template, send_from_directory, session
import pandas as pd
from sklearn.externals import joblib
from sklearn.ensemble import RandomForestClassifier
import numpy as np
import urllib.request
import json
import random
import openpyxl
from geopy.geocoders import Nominatim
from flask_session import Session
import csv
from csv import writer

app = Flask(__name__)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SECRET_KEY'] = '123456'
Session(app)


# Using Render_Template it Directly Routes from Templates Folder

@app.route('/')
def general():
    return render_template('signup.html')


@app.route('/index.html')
def index():
    return render_template('index.html')


@app.route('/index2.html')
def index2():
    return render_template('index2.html')


def root():
    return render_template('index.html')


@app.route('/work.html')
def work():
    return render_template('work.html')


@app.route('/addreport.html')
def addreport():
    return render_template('addreport.html')


@app.route('/addreport', methods=['POST'])
def addreport1():
    location = request.form['Location']
    timestamp = request.form['timestamp']
    crimetype = request.form['crime-type']
    des = request.form['des']

    geolocator = Nominatim()
    loc = geolocator.geocode(location, timeout=None)

    if loc == None:
        lat = 13.3275
        log = 80.1781
    else:
        lat = loc.latitude
        log = loc.longitude

    data = [session['username'], location, timestamp]
    if crimetype == "roberry":
        data += [1, 0, 0, 0, 0, 0]
    elif crimetype == "gambling":
        data += [0, 1, 0, 0, 0, 0]
    elif crimetype == "accident":
        data += [0, 0, 1, 0, 0, 0]
    elif crimetype == "violence":
        data += [0, 0, 0, 1, 0, 0]
    elif crimetype == "murder":
        data += [0, 0, 0, 0, 1, 0]
    else:
        data += [0, 0, 0, 0, 0, 1]

    data += [lat, log, des, crimetype, 0]

    workbook = openpyxl.load_workbook('model/user_reports.xlsx')

    # Select the sheet you want to add a row to
    worksheet = workbook['Sheet1']

    # Add the new row to the sheet
    worksheet.append(data)

    # Save the changes to the Excel file
    workbook.save('model/user_reports.xlsx')

    print(location, timestamp, crimetype, des, lat, log)

    return render_template('mainpage.html', message="Welcome " + session['username'],
                           text="Report has passed our initial review and is now ready for Admin verification.")


@app.route('/analysis.html')
def analysis():
    csv_file_path = 'data.csv'

    df = pd.read_csv(csv_file_path)
    num_rows = len(df.index)
    rand_rows = random.sample(range(num_rows), 100)
    df_rand_rows = df.iloc[rand_rows, :]
    counts = df_rand_rows.sum(axis=0)

    Data = [
        ["Robbery", counts['act379']],
        ["Gambling", counts['act13']],
        ["Accident", counts['act279']],
        ["Violence", counts['act323']],
        ["Murder", counts['act302']],
        ["Kidnapping", counts['act363']]
    ]

    return render_template('analysis.html', data=Data)


@app.route('/about.html')
def about():
    return render_template('about.html')


@app.route('/about2.html')
def about2():
    return render_template('about2.html')

@app.route('/admin.html')
def admin():
    List = []
    head = []
    workbook = openpyxl.load_workbook('model/user_reports.xlsx')
    worksheet = workbook.active
    for row_number, row in enumerate(worksheet.iter_rows()):
        tem = [cell.value for cell in row]
        if row_number == 0:
            head += ["ID", tem[0], tem[1], tem[2], tem[12], tem[11]]
        else:
            if tem[13] == 0:
                List.append([row_number, tem[0], tem[1], tem[2], tem[12], tem[11]])
    return render_template('admin.html', message="Verification Successful....", head=head, List=List)

@app.route('/verified.html')
def verified():
    List = []
    head = []
    workbook = openpyxl.load_workbook('model/user_reports.xlsx')
    worksheet = workbook.active
    for row_number, row in enumerate(worksheet.iter_rows()):
        tem = [cell.value for cell in row]
        if row_number == 0:
            head += ["ID", tem[0], tem[1], tem[2], tem[12], tem[11]]
        else:
            if tem[13] == 1:
                List.append([row_number, tem[0], tem[1], tem[2], tem[12], tem[11]])
    return render_template('verified.html', message="Verified Reports....", head=head, List=List)


@app.route('/report.html')
def report():
    return render_template('report.html')


@app.route('/logout.html')
def logout():
    session.pop('username', None)
    session.pop('logged_in', None)
    return render_template('signup.html', message="Logout Successful!!!")


@app.route('/signup.html')
def signup():
    return render_template('signup.html')


@app.route('/main.html')
def main():
    if 'logged_in' in session:
        return render_template('mainpage.html', message="Welcome " + session['username'])
    else:
        return render_template('signup.html', message="Session Expired....Login Again!!!")


@app.route('/verify', methods=['POST'])
def verify():
    ID = request.form['ID']

    workbook = openpyxl.load_workbook('model/user_reports.xlsx')

    sheet = workbook['Sheet1']

    sheet.cell(row=int(ID) + 1, column=14, value=1)

    # Save the workbook
    workbook.save('model/user_reports.xlsx')

    worksheet = workbook.active

    row_data = [cell.value for cell in worksheet[int(ID) + 1]]

    data = row_data[2:11]
    print(data)
    file = open('data.csv', 'a+')

    # writing the data into the file
    with file:
        write = csv.writer(file)
        write.writerow(data)

    List = []
    head = []
    workbook = openpyxl.load_workbook('model/user_reports.xlsx')
    worksheet = workbook.active
    for row_number, row in enumerate(worksheet.iter_rows()):
        tem = [cell.value for cell in row]
        if row_number == 0:
            head += ["ID", tem[0], tem[1], tem[2], tem[12], tem[11]]
        else:
            if tem[13] == 0:
                List.append([row_number, tem[0], tem[1], tem[2], tem[12], tem[11]])
    return render_template('admin.html', message="Verification Successful....", head=head, List=List)


@app.route('/signup', methods=['POST'])
def signupPro():
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']

    # Do something with the form data here (e.g. store it in a database)
    print(username, email, password)

    workbook = openpyxl.load_workbook('model/userdata.xlsx')

    # Select the sheet you want to add a row to
    worksheet = workbook['Sheet1']

    # Define the data for the new row
    new_row = [username, email, password]

    # Add the new row to the sheet
    worksheet.append(new_row)

    # Save the changes to the Excel file
    workbook.save('model/userdata.xlsx')

    return render_template('signup.html', message="Account Created Successfully....Login Now!!!")


@app.route('/signin', methods=['POST'])
def signinPro():
    email = request.form['email1']
    password = request.form['password1']

    if email == "admintools@gmail.com" and password == "admin":
        List = []
        head = []
        workbook = openpyxl.load_workbook('model/user_reports.xlsx')
        worksheet = workbook.active
        for row_number, row in enumerate(worksheet.iter_rows()):
            tem = [cell.value for cell in row]
            if row_number == 0:
                head += ["ID", tem[0], tem[1], tem[2], tem[12], tem[11]]
            else:
                if tem[13] == 0:
                    List.append([row_number, tem[0], tem[1], tem[2], tem[12], tem[11]])
        return render_template('admin.html', List=List, head=head)

    wb = openpyxl.load_workbook('model/userdata.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Iterate over each row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Access each cell value in the row
        dbname = row[0]
        dbmail = row[1]
        dbpass = row[2]
        if email == dbmail and password == dbpass:
            session['logged_in'] = True
            session['username'] = dbname
            return render_template('mainpage.html', message="Welcome " + dbname)

    return render_template('signup.html', message="Invalid Login Details!!!")


@app.route('/result.html', methods=['POST'])
def predict():
    rfc = joblib.load('model/rf_model')
    print('model loaded')

    if request.method == 'POST':

        address = request.form['Location']
        geolocator = Nominatim()
        location = geolocator.geocode(address, timeout=None)
        if location == None:
            return render_template('result.html', prediction="Predicted crime : Act 279-Accident")
        print(location.address)
        lat = [location.latitude]
        log = [location.longitude]
        latlong = pd.DataFrame({'latitude': lat, 'longitude': log})
        print(latlong)

        DT = request.form['timestamp']
        latlong['timestamp'] = DT
        data = latlong
        cols = data.columns.tolist()
        cols = cols[-1:] + cols[:-1]
        data = data[cols]

        data['timestamp'] = pd.to_datetime(data['timestamp'].astype(str), errors='coerce')
        data['timestamp'] = pd.to_datetime(data['timestamp'], format='%d/%m/%Y %H:%M:%S')
        column_1 = data.ix[:, 0]
        DT = pd.DataFrame({"year": column_1.dt.year,
                           "month": column_1.dt.month,
                           "day": column_1.dt.day,
                           "hour": column_1.dt.hour,
                           "dayofyear": column_1.dt.dayofyear,
                           "week": column_1.dt.week,
                           "weekofyear": column_1.dt.weekofyear,
                           "dayofweek": column_1.dt.dayofweek,
                           "weekday": column_1.dt.weekday,
                           "quarter": column_1.dt.quarter,
                           })
        data = data.drop('timestamp', axis=1)
        final = pd.concat([DT, data], axis=1)
        X = final.iloc[:, [1, 2, 3, 4, 6, 10, 11]].values
        my_prediction = rfc.predict(X)
        if my_prediction[0][0] == 1:
            my_prediction = 'Predicted crime : Act 379-Robbery'
        elif my_prediction[0][1] == 1:
            my_prediction = 'Predicted crime : Act 13-Gambling'
        elif my_prediction[0][2] == 1:
            my_prediction = 'Predicted crime : Act 279-Accident'
        elif my_prediction[0][3] == 1:
            my_prediction = 'Predicted crime : Act 323-Violence'
        elif my_prediction[0][4] == 1:
            my_prediction = 'Predicted crime : Act 302-Murder'
        elif my_prediction[0][5] == 1:
            my_prediction = 'Predicted crime : Act 363-kidnapping'
        else:
            my_prediction = 'Place is safe no crime expected at that timestamp.'

    return render_template('result.html', prediction=my_prediction)


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8080, threaded=True)
